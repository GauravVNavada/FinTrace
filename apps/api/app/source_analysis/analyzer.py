import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from app.financial_investigations.files import UploadValidationError
from app.source_analysis.schemas import ColumnProfile, InferredType


class AnalysisDocument:
    def __init__(
        self,
        headers: list[str],
        sample_rows: list[dict[str, str | None]],
        profiles: list[ColumnProfile],
        row_count: int,
        column_count: int,
        rows: list[list[str | None]],
    ) -> None:
        self.headers = headers
        self.sample_rows = sample_rows
        self.profiles = profiles
        self.row_count = row_count
        self.column_count = column_count
        self.rows = rows


def analyze_content(
    filename: str,
    content: bytes,
    max_rows: int,
    max_columns: int,
    *,
    truncate: bool = False,
) -> AnalysisDocument:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        headers, rows = _read_csv(content, max_rows, max_columns, truncate=truncate)
    elif extension == ".xlsx":
        headers, rows = _read_xlsx(content, max_rows, max_columns, truncate=truncate)
    else:
        raise UploadValidationError("Only CSV and XLSX files can be analyzed")
    profiles = [
        _profile_column(header, [row[index] for row in rows])
        for index, header in enumerate(headers)
    ]
    sample_rows = [
        dict(zip(headers, [_display(value) for value in row], strict=True)) for row in rows[:20]
    ]
    return AnalysisDocument(headers, sample_rows, profiles, len(rows), len(headers), rows)


def _read_csv(
    content: bytes, max_rows: int, max_columns: int, *, truncate: bool
) -> tuple[list[str], list[list[str | None]]]:
    try:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), strict=True)
        raw_header = next(reader, None)
        if not raw_header:
            raise UploadValidationError("The CSV must contain a header row")
        headers = [value.strip() for value in raw_header]
        rows: list[list[str | None]] = []
        for raw_row in reader:
            if len(raw_row) > max_columns:
                raise UploadValidationError("The CSV contains too many columns")
            if len(raw_row) > len(headers):
                raise UploadValidationError("The CSV row contains more columns than the header")
            padded = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
            if len(rows) >= max_rows and not truncate:
                raise UploadValidationError(
                    "The CSV contains more rows than the configured analysis limit"
                )
            rows.append(padded)
            if len(rows) >= max_rows and truncate:
                break
        return headers, rows
    except UnicodeDecodeError as error:
        raise UploadValidationError("The CSV must be UTF-8 encoded") from error
    except csv.Error as error:
        raise UploadValidationError("The CSV is malformed") from error


def _read_xlsx(
    content: bytes, max_rows: int, max_columns: int, *, truncate: bool
) -> tuple[list[str], list[list[str | None]]]:
    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        sheet = workbook.active
        if sheet is None:
            raise UploadValidationError("The XLSX does not contain an active worksheet")
        iterator = sheet.iter_rows(values_only=True)
        raw_header = next(iterator, None)
        if not raw_header:
            raise UploadValidationError("The XLSX must contain a header row")
        headers = ["" if value is None else str(value).strip() for value in raw_header]
        rows: list[list[str | None]] = []
        for raw_row in iterator:
            if len(raw_row) > max_columns:
                raise UploadValidationError("The XLSX contains too many columns")
            if len(rows) >= max_rows and not truncate:
                raise UploadValidationError(
                    "The XLSX contains more rows than the configured analysis limit"
                )
            rows.append(
                [None if value is None else _display(value) for value in raw_row[: len(headers)]]
                + [None] * max(0, len(headers) - len(raw_row))
            )
            if len(rows) >= max_rows and truncate:
                break
        return headers, rows
    except UploadValidationError:
        raise
    except Exception as error:
        raise UploadValidationError("The XLSX is malformed or cannot be read") from error
    finally:
        if workbook is not None:
            workbook.close()


def _profile_column(name: str, values: list[str | None]) -> ColumnProfile:
    clean = [value.strip() for value in values if value is not None and value.strip()]
    inferred_type = _infer_type(clean)
    numeric_values: list[Decimal] = []
    if inferred_type in {InferredType.INTEGER, InferredType.DECIMAL}:
        numeric_values = [_decimal(value) for value in clean]
    return ColumnProfile(
        name=name,
        inferred_type=inferred_type,
        non_empty_count=len(clean),
        unique_count=len(set(clean)),
        sample_values=clean[:5],
        min_value=str(min(numeric_values)) if numeric_values else (min(clean) if clean else None),
        max_value=str(max(numeric_values)) if numeric_values else (max(clean) if clean else None),
    )


def _infer_type(values: list[str]) -> InferredType:
    if not values:
        return InferredType.EMPTY
    lowered = {value.casefold() for value in values}
    if lowered <= {"true", "false", "yes", "no"}:
        return InferredType.BOOLEAN
    if all(_is_integer(value) for value in values):
        return InferredType.INTEGER
    if all(_is_decimal(value) for value in values):
        return InferredType.DECIMAL
    if all(_is_date(value) for value in values):
        return InferredType.DATE
    return InferredType.STRING


def _is_integer(value: str) -> bool:
    try:
        return Decimal(value).as_tuple().exponent == 0
    except InvalidOperation:
        return False


def _is_decimal(value: str) -> bool:
    try:
        Decimal(value)
        return True
    except InvalidOperation:
        return False


def _is_date(value: str) -> bool:
    try:
        date.fromisoformat(value[:10])
        if "T" in value or " " in value:
            datetime.fromisoformat(value)
        return True
    except ValueError:
        return False


def _decimal(value: str) -> Decimal:
    return Decimal(value)


def _display(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)
