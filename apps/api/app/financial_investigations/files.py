import csv
import hashlib
import io
import zipfile
from dataclasses import dataclass
from pathlib import Path
from uuid import uuid4

from app.core.config import get_settings


class UploadValidationError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class FileInspection:
    extension: str
    mime_type: str
    size_bytes: int
    row_count: int
    column_count: int
    sha256: str


def inspect_upload(
    filename: str | None, content_type: str | None, content: bytes
) -> FileInspection:
    safe_name = Path(filename or "").name
    extension = Path(safe_name).suffix.lower()
    if safe_name in {"", ".", ".."} or extension not in {".csv", ".xlsx"}:
        raise UploadValidationError("Only CSV and XLSX files are supported")
    settings = get_settings()
    size = len(content)
    if size == 0:
        raise UploadValidationError("The uploaded file is empty")
    if size > settings.max_upload_bytes:
        raise UploadValidationError(
            f"The uploaded file exceeds the {settings.max_upload_bytes} byte limit"
        )
    if extension == ".csv":
        rows, columns = _inspect_csv(content, settings.max_upload_rows, settings.max_upload_columns)
        mime_type = "text/csv"
    else:
        rows, columns = _inspect_xlsx(
            content, settings.max_upload_rows, settings.max_upload_columns
        )
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if content_type and content_type not in {
        "text/csv",
        "text/plain",
        "application/csv",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }:
        raise UploadValidationError("The uploaded content type is not supported")
    return FileInspection(
        extension, mime_type, size, rows, columns, hashlib.sha256(content).hexdigest()
    )


def store_upload(content: bytes, inspection: FileInspection) -> str:
    settings = get_settings()
    root = (Path(__file__).resolve().parents[2] / settings.upload_directory).resolve()
    root.mkdir(parents=True, exist_ok=True)
    stored_name = f"{uuid4().hex}{inspection.extension}"
    target = (root / stored_name).resolve()
    if root not in target.parents:
        raise UploadValidationError("Unable to create a safe storage reference")
    target.write_bytes(content)
    return str(target.relative_to(Path(__file__).resolve().parents[2]))


def remove_upload(storage_reference: str) -> None:
    root = (Path(__file__).resolve().parents[2] / get_settings().upload_directory).resolve()
    target = (Path(__file__).resolve().parents[2] / storage_reference).resolve()
    if root not in target.parents:
        return
    target.unlink(missing_ok=True)


def read_upload(storage_reference: str) -> tuple[str, bytes]:
    """Read a stored upload only through a server-generated reference."""
    root = (Path(__file__).resolve().parents[2] / get_settings().upload_directory).resolve()
    target = (Path(__file__).resolve().parents[2] / storage_reference).resolve()
    if root not in target.parents or not target.is_file():
        raise UploadValidationError("The stored source file is unavailable")
    content = target.read_bytes()
    if len(content) > get_settings().max_upload_bytes:
        raise UploadValidationError("The stored source file exceeds the configured limit")
    return target.name, content


def _inspect_csv(content: bytes, max_rows: int, max_columns: int) -> tuple[int, int]:
    try:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), strict=True)
        header = next(reader, None)
        if not header or not any(cell.strip() for cell in header):
            raise UploadValidationError("The CSV must contain a header row")
        columns = len(header)
        _validate_header(header, max_columns)
        rows = 0
        for row in reader:
            if len(row) > max_columns:
                raise UploadValidationError("The CSV contains too many columns")
            rows += 1
            if rows > max_rows:
                raise UploadValidationError("The CSV contains too many rows")
        return rows, columns
    except UnicodeDecodeError as error:
        raise UploadValidationError("The CSV must be UTF-8 encoded") from error
    except csv.Error as error:
        raise UploadValidationError("The CSV is malformed") from error


def _inspect_xlsx(content: bytes, max_rows: int, max_columns: int) -> tuple[int, int]:
    if not zipfile.is_zipfile(io.BytesIO(content)):
        raise UploadValidationError("The XLSX file is malformed")
    settings = get_settings()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > settings.max_upload_zip_members:
                raise UploadValidationError("The XLSX archive contains too many members")
            uncompressed_size = sum(member.file_size for member in members)
            if uncompressed_size > settings.max_upload_uncompressed_bytes:
                raise UploadValidationError(
                    "The XLSX uncompressed size exceeds the configured limit"
                )
            if any(member.file_size > settings.max_upload_uncompressed_bytes for member in members):
                raise UploadValidationError("The XLSX contains an oversized archive member")
    except UploadValidationError:
        raise
    except (OSError, zipfile.BadZipFile) as error:
        raise UploadValidationError("The XLSX archive is malformed") from error
    workbook = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        sheet = workbook.active
        if sheet is None:
            raise UploadValidationError("The XLSX does not contain an active worksheet")
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header or not any(value is not None and str(value).strip() for value in header):
            raise UploadValidationError("The XLSX must contain a header row")
        normalized_header = ["" if value is None else str(value) for value in header]
        columns = len(normalized_header)
        _validate_header(normalized_header, max_columns)
        rows = 0
        for row in iterator:
            if len(row) > max_columns:
                raise UploadValidationError("The XLSX contains too many columns")
            rows += 1
            if rows > max_rows:
                raise UploadValidationError("The XLSX contains too many rows")
        return rows, columns
    except UploadValidationError:
        raise
    except Exception as error:
        raise UploadValidationError("The XLSX is malformed or cannot be read") from error
    finally:
        if workbook is not None:
            workbook.close()


def _validate_header(header: list[str], max_columns: int) -> None:
    if len(header) > max_columns:
        raise UploadValidationError("The file contains too many columns")
    normalized = [value.strip().casefold() for value in header]
    if any(not value for value in normalized) or len(set(normalized)) != len(normalized):
        raise UploadValidationError("The file header must contain unique, non-empty column names")
